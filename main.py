import calendar
import datetime
import openpyxl
import locale

from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.styles.borders import Border, Side, BORDER_THIN


class CalendarExcelGenerator:

    def __init__(self):
        self.book = openpyxl.Workbook()
        self.column_width = 30
        self.headerFill = PatternFill(start_color='FF7d7777',
                    end_color='FF7d7777',
                    fill_type='solid')
        self.dateFill = PatternFill(start_color='FF6d87ab',
                    end_color='FF6d87ab',
                    fill_type='solid')
        self.cells = ["a", "b", "c", "d", "e", "f", "g", "h"]
        self.weekdays = ["Тип", "Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота", "Воскресенье"]
        self.mounts = ["Январь", "Февраль", "Март", "Апрель", "Май", "Июнь", "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"]
        self.year = 2024 # datetime.datetime.now().year --> current year


    def createCellParams(self, sheet, cellIdent, fontSize, color):
        headerCell = sheet[cellIdent] 
        headerCell.alignment = Alignment(horizontal='center', vertical='center')
        #headerCell.verticalAlignment = Alignment(verticalhorizontal='center')
        headerCell.font = Font(size=fontSize)
        headerCell.fill = color

    def generateMount(self, indexMount):
        sheet = self.book.create_sheet(self.mounts[indexMount])
        rowColumn = 1
        # заголовок
        sheet.merge_cells(start_row=rowColumn, start_column=1, end_row=rowColumn, end_column=len(self.weekdays) + 1)
        self.createCellParams(sheet, "A"+str(rowColumn), 20, self.headerFill)
        sheet["A"+str(rowColumn)] = self.mounts[indexMount] + " " + str(self.year) + " года"
        rowColumn += 1

        # дни недели
        for i in range(0, 8):
            self.createCellParams(sheet, self.cells[i] + str(rowColumn), 16, self.headerFill)
            sheet.column_dimensions[get_column_letter(i+1)].width = self.column_width
            sheet[self.cells[i] + str(rowColumn)] = self.weekdays[i]
        self.createCellParams(sheet, "I" + str(rowColumn), 16, self.headerFill)
        rowColumn += 1

        # ячейки
        for week in calendar.monthcalendar(self.year, indexMount + 1):
            index = 0
            self.createCellParams(sheet, self.cells[index] + str(rowColumn), 14, self.dateFill)
            self.createCellParams(sheet, self.cells[index] + str(rowColumn + 1), 14, self.dateFill)
            self.createCellParams(sheet, self.cells[index] + str(rowColumn + 2), 14, self.dateFill)
            sheet["A" + str(rowColumn + 1)]  = "План"
            sheet["A" + str(rowColumn + 2)]  = "Результат"
            index += 1
            for day in week:
                self.createCellParams(sheet, self.cells[index] + str(rowColumn), 14, self.dateFill)
                if(day != 0):
                    date = datetime.datetime(self.year, indexMount + 1, day)
                    sheet[self.cells[date.weekday() + 1] + str(rowColumn)]  = date.strftime("%d.%m")
                index += 1
            sheet.column_dimensions[get_column_letter(index + 1)].width = self.column_width
            self.createCellParams(sheet, "I" + str(rowColumn), 14, self.dateFill)
            sheet["I" + str(rowColumn)]  = "Объём за неделю"
            rowColumn += 1
            sheet.row_dimensions[rowColumn].height = 70
            sheet.column_dimensions[get_column_letter(index + 1)].width = self.column_width
            rowColumn += 1
            sheet.row_dimensions[rowColumn].height = 70
            sheet.merge_cells(start_row=rowColumn-1, start_column=9, end_row=rowColumn, end_column=9)
            rowColumn += 1

        # границы
        thin_border = Border(left=Side(border_style=BORDER_THIN, color='00000000'), right=Side(border_style=BORDER_THIN, color='00000000'), top=Side(border_style=BORDER_THIN, color='00000000'), bottom=Side(border_style=BORDER_THIN, color='00000000'))
        for i in range(1, 10):
            for j in range (1, rowColumn):
                sheet.cell(row=j, column=i).border = thin_border

    def generateFile(self):
        for i in range(0, len(self.mounts)):
           self.generateMount(i) 
        self.book.remove(self.book["Sheet"])
        self.book.save('calendar_training.xlsx')

def main():
    locale.setlocale(locale.LC_TIME, 'ru_RU.UTF-8')
    generator = CalendarExcelGenerator()
    generator.generateFile()

if __name__ == "__main__":
    main()